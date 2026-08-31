#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Informe detallado de diferencias de despacho por picker.

Toma el archivo INFORME_DIFERENCIAS.xlsx (Hoja1 = diferencias de inventario,
Hoja2 = despachos con el picker responsable de cada entrega), cruza ambas hojas
por el numero de entrega y genera un libro de Excel con:

  * Resumen         : indicadores generales del periodo.
  * Ranking Pickers : quien genera mas diferencias y cuanto cuesta.
  * Picker x Tienda : en que tienda tiene diferencias cada picker y su valor.
  * Detalle         : linea a linea, con el picker responsable.
  * Ranking Tiendas : tiendas con mayor impacto economico.
  * Top Materiales  : materiales que mas se pierden.
  * Sin Picker      : diferencias que no se pudieron atribuir (trazabilidad).

Uso:
    python informe_diferencias.py INFORME_DIFERENCIAS.xlsx
    python informe_diferencias.py entrada.xlsx -o informe_pickers.xlsx
    python informe_diferencias.py entrada.xlsx --estado todos --metrica neto
"""

from __future__ import annotations

import argparse
import re
import sys
import unicodedata
from dataclasses import dataclass
from pathlib import Path

import pandas as pd

# --------------------------------------------------------------------------- #
# Utilidades de normalizacion
# --------------------------------------------------------------------------- #

FORMATO_PESOS = '"$"#,##0;[Red]-"$"#,##0'


def normaliza(texto) -> str:
    """Minusculas, sin tildes, sin saltos de linea y sin espacios repetidos."""
    if texto is None or (isinstance(texto, float) and pd.isna(texto)):
        return ""
    texto = str(texto).replace("\n", " ").replace("\r", " ")
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(c for c in texto if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", texto).strip().lower()


def busca_columna(df: pd.DataFrame, alias: list[str], obligatoria: bool = True):
    """Devuelve el nombre real de la primera columna que coincide con un alias."""
    mapa = {normaliza(c): c for c in df.columns}
    for a in alias:
        if a in mapa:
            return mapa[a]
    for a in alias:  # segundo intento: coincidencia parcial
        for norm, real in mapa.items():
            if a in norm:
                return real
    if obligatoria:
        raise KeyError(
            f"No se encontro ninguna columna para {alias!r}. "
            f"Columnas disponibles: {list(df.columns)}"
        )
    return None


def a_numero(serie: pd.Series) -> pd.Series:
    """Convierte a numero tolerando '$', puntos de miles, comas y guiones."""
    if pd.api.types.is_numeric_dtype(serie):
        return pd.to_numeric(serie, errors="coerce")
    txt = (
        serie.astype(str)
        .str.replace(r"[^\d,.\-]", "", regex=True)
        .str.replace(".", "", regex=False)
        .str.replace(",", ".", regex=False)
        .replace({"": None, "-": None})
    )
    return pd.to_numeric(txt, errors="coerce")


def limpia_texto(serie: pd.Series) -> pd.Series:
    return (
        serie.astype("string")
        .str.replace(r"\s+", " ", regex=True)
        .str.strip()
        .str.upper()
        .replace({"": pd.NA, "NAN": pd.NA, "#VALUE!": pd.NA, "#N/A": pd.NA})
    )


def detecta_fila_encabezado(ruta: Path, hoja, pistas: list[str], max_filas: int = 12) -> int:
    """Busca la fila que funciona como encabezado (la que mas pistas contiene)."""
    crudo = pd.read_excel(ruta, sheet_name=hoja, header=None, nrows=max_filas)
    mejor, mejor_puntaje = 0, -1
    for i in range(len(crudo)):
        celdas = [normaliza(v) for v in crudo.iloc[i].tolist()]
        puntaje = sum(any(p in c for c in celdas) for p in pistas)
        if puntaje > mejor_puntaje:
            mejor, mejor_puntaje = i, puntaje
    return mejor


# --------------------------------------------------------------------------- #
# Carga y cruce
# --------------------------------------------------------------------------- #

@dataclass
class Datos:
    detalle: pd.DataFrame          # diferencias con picker (cuando se pudo cruzar)
    despachos: pd.DataFrame        # hoja de despachos normalizada
    sin_picker: pd.DataFrame       # diferencias sin picker identificado
    cobertura: dict                # metricas del cruce


def carga_diferencias(ruta: Path, hoja) -> pd.DataFrame:
    pistas = ["nombre tienda", "codigo material", "estado", "total", "cuadrante"]
    fila = detecta_fila_encabezado(ruta, hoja, pistas)
    df = pd.read_excel(ruta, sheet_name=hoja, header=fila)
    df = df.dropna(how="all")

    col = {
        "fecha": busca_columna(df, ["descarga inventarios (fecha/hora)", "fecha"]),
        "ceco": busca_columna(df, ["ceco2", "ceco"]),
        "tienda": busca_columna(df, ["nombre tienda", "tienda"]),
        "material": busca_columna(df, ["codigo material", "material"]),
        "cuadrante": busca_columna(df, ["cuadrante"]),
        "descripcion": busca_columna(df, ["descripcion material", "descripcion"]),
        "precio": busca_columna(df, ["precio unitario", "precio"]),
        "unidades": busca_columna(df, ["unids. reporte", "unidades", "unids"]),
        "valor": busca_columna(df, ["$ total", "total"]),
        "estado": busca_columna(df, ["estado"]),
        "tipo_dif": busca_columna(df, ["falt/sob", "faltante"]),
        "causal": busca_columna(df, ["entrega inexistente/sin ingreso", "causal"], False),
        "llave": busca_columna(df, ["comprobacion duplicados", "cruce senuelos", "llave"], False),
        "documento": busca_columna(df, ["documento"], False),
        "area": busca_columna(df, ["area"], False),
        "gv": busca_columna(df, ["gv"], False),
        "jdz": busca_columna(df, ["jdz"], False),
        "atiende": busca_columna(df, ["atiende"], False),
    }

    out = pd.DataFrame(index=df.index)
    out["FECHA"] = pd.to_datetime(df[col["fecha"]], errors="coerce")
    out["CECO"] = limpia_texto(df[col["ceco"]])
    out["TIENDA"] = limpia_texto(df[col["tienda"]])
    out["CUADRANTE"] = limpia_texto(df[col["cuadrante"]])
    out["CODIGO MATERIAL"] = a_numero(df[col["material"]]).astype("Int64")
    out["DESCRIPCION MATERIAL"] = limpia_texto(df[col["descripcion"]])
    out["PRECIO UNITARIO"] = a_numero(df[col["precio"]])
    out["UNIDADES"] = a_numero(df[col["unidades"]])
    out["VALOR"] = a_numero(df[col["valor"]])
    out["ESTADO"] = limpia_texto(df[col["estado"]])
    out["TIPO DIFERENCIA"] = limpia_texto(df[col["tipo_dif"]])
    for destino, clave in [("CAUSAL", "causal"), ("DOCUMENTO", "documento"),
                           ("AREA", "area"), ("GV", "gv"), ("JDZ", "jdz"),
                           ("ATIENDE", "atiende")]:
        out[destino] = limpia_texto(df[col[clave]]) if col[clave] else pd.NA

    # El valor puede venir vacio: se reconstruye con precio x unidades.
    reconstruido = out["PRECIO UNITARIO"] * out["UNIDADES"]
    out["VALOR"] = out["VALOR"].fillna(reconstruido)

    # Si no viene marcado, se deduce del signo del valor.
    out["TIPO DIFERENCIA"] = out["TIPO DIFERENCIA"].fillna(
        pd.Series(["FALTANTE" if v < 0 else "SOBRANTE" for v in out["VALOR"].fillna(0)],
                  index=out.index, dtype="string")
    )

    # Numero de entrega: la llave concatena CECO(10) + entrega(10) + material(8).
    llave = df[col["llave"]].astype("string") if col["llave"] else pd.Series(pd.NA, index=df.index, dtype="string")
    entrega = llave.str.extract(r"^.{10}(\d{10})")[0]
    entrega = entrega.fillna(llave.str.extract(r"(98\d{8})")[0])
    out["N ENTREGA"] = entrega

    out = out[out["TIENDA"].notna() | out["VALOR"].notna()]
    return out.reset_index(drop=True)


def carga_despachos(ruta: Path, hoja) -> pd.DataFrame:
    pistas = ["picker", "entrega", "tienda", "ruta", "cajas"]
    fila = detecta_fila_encabezado(ruta, hoja, pistas)
    df = pd.read_excel(ruta, sheet_name=hoja, header=fila).dropna(how="all")

    c_ent = busca_columna(df, ["n° entrega", "no entrega", "entrega"])
    c_pick = busca_columna(df, ["picker"])
    c_ruta = busca_columna(df, ["ruta"], False)
    c_tienda = busca_columna(df, ["tienda"], False)
    c_cuad = busca_columna(df, ["cuadrante"], False)
    c_cajas = busca_columna(df, ["suma de cajas", "cajas"], False)
    c_fecha = busca_columna(df, ["fecha"], False)

    out = pd.DataFrame(index=df.index)
    out["N ENTREGA"] = (
        a_numero(df[c_ent]).astype("Int64").astype("string").replace({"<NA>": pd.NA})
    )
    out["PICKER"] = limpia_texto(df[c_pick]).fillna("SIN PICKER")
    out["RUTA"] = limpia_texto(df[c_ruta]) if c_ruta else pd.NA
    out["TIENDA DESPACHO"] = limpia_texto(df[c_tienda]) if c_tienda else pd.NA
    out["CUADRANTE DESPACHO"] = limpia_texto(df[c_cuad]) if c_cuad else pd.NA
    out["CAJAS"] = a_numero(df[c_cajas]) if c_cajas else 0.0
    out["FECHA DESPACHO"] = pd.to_datetime(df[c_fecha], errors="coerce") if c_fecha else pd.NaT

    out = out[out["N ENTREGA"].notna()]
    # Una entrega puede repetirse: se consolida en una sola fila.
    out = (
        out.sort_values("FECHA DESPACHO")
        .groupby("N ENTREGA", as_index=False)
        .agg({
            "PICKER": "first", "RUTA": "first", "TIENDA DESPACHO": "first",
            "CUADRANTE DESPACHO": "first", "CAJAS": "sum", "FECHA DESPACHO": "first",
        })
    )
    return out


def cruza(diferencias: pd.DataFrame, despachos: pd.DataFrame) -> Datos:
    unido = diferencias.merge(despachos, on="N ENTREGA", how="left")
    con_picker = unido[unido["PICKER"].notna()].copy()
    sin_picker = unido[unido["PICKER"].isna()].copy()

    cobertura = {
        "Lineas de diferencia": len(unido),
        "Lineas con picker identificado": len(con_picker),
        "Lineas sin picker identificado": len(sin_picker),
        "% lineas atribuidas": round(100 * len(con_picker) / max(len(unido), 1), 1),
        "Entregas en hoja de despacho": len(despachos),
        "Pickers identificados": int(con_picker["PICKER"].nunique()),
    }
    return Datos(con_picker, despachos, sin_picker, cobertura)


# --------------------------------------------------------------------------- #
# Metricas
# --------------------------------------------------------------------------- #

def agrega(df: pd.DataFrame, llaves: list[str]) -> pd.DataFrame:
    """Metricas economicas estandar para cualquier corte (picker, tienda, etc.)."""
    if df.empty:
        return pd.DataFrame(columns=llaves + [
            "LINEAS", "UNIDADES FALTANTES", "UNIDADES SOBRANTES",
            "VALOR FALTANTES", "VALOR SOBRANTES", "IMPACTO NETO", "IMPACTO ABSOLUTO",
        ])
    d = df.copy()
    d["_falt"] = d["VALOR"].where(d["VALOR"] < 0, 0).abs()
    d["_sobr"] = d["VALOR"].where(d["VALOR"] > 0, 0)
    d["_uf"] = d["UNIDADES"].where(d["UNIDADES"] < 0, 0).abs()
    d["_us"] = d["UNIDADES"].where(d["UNIDADES"] > 0, 0)

    g = d.groupby(llaves, dropna=False).agg(
        LINEAS=("VALOR", "size"),
        **{"UNIDADES FALTANTES": ("_uf", "sum"),
           "UNIDADES SOBRANTES": ("_us", "sum"),
           "VALOR FALTANTES": ("_falt", "sum"),
           "VALOR SOBRANTES": ("_sobr", "sum")},
    ).reset_index()
    g["IMPACTO NETO"] = g["VALOR SOBRANTES"] - g["VALOR FALTANTES"]
    g["IMPACTO ABSOLUTO"] = g["VALOR FALTANTES"] + g["VALOR SOBRANTES"]
    return g


def ranking_pickers(datos: Datos, metrica: str) -> pd.DataFrame:
    df = datos.detalle
    r = agrega(df, ["PICKER"])
    if r.empty:
        return r

    extra = df.groupby("PICKER").agg(
        **{"TIENDAS CON DIFERENCIA": ("TIENDA", "nunique"),
           "ENTREGAS CON DIFERENCIA": ("N ENTREGA", "nunique"),
           "MATERIALES DISTINTOS": ("CODIGO MATERIAL", "nunique")}
    ).reset_index()
    r = r.merge(extra, on="PICKER", how="left")

    # Exposicion del picker: todo lo que despacho en el periodo.
    exp = datos.despachos.groupby("PICKER").agg(
        **{"ENTREGAS DESPACHADAS": ("N ENTREGA", "nunique"),
           "CAJAS DESPACHADAS": ("CAJAS", "sum")}
    ).reset_index()
    r = r.merge(exp, on="PICKER", how="left")
    r[["ENTREGAS DESPACHADAS", "CAJAS DESPACHADAS"]] = (
        r[["ENTREGAS DESPACHADAS", "CAJAS DESPACHADAS"]].fillna(0)
    )

    r["% ENTREGAS CON DIFERENCIA"] = (
        100 * r["ENTREGAS CON DIFERENCIA"] / r["ENTREGAS DESPACHADAS"].replace(0, pd.NA)
    ).round(1)
    r["COSTO POR CAJA"] = (
        r["VALOR FALTANTES"] / r["CAJAS DESPACHADAS"].replace(0, pd.NA)
    ).round(0)
    r["COSTO POR ENTREGA DESPACHADA"] = (
        r["VALOR FALTANTES"] / r["ENTREGAS DESPACHADAS"].replace(0, pd.NA)
    ).round(0)

    columna_orden = {
        "faltante": "VALOR FALTANTES",
        "neto": "IMPACTO NETO",
        "absoluto": "IMPACTO ABSOLUTO",
    }[metrica]
    ascendente = metrica == "neto"
    r = r.sort_values(columna_orden, ascending=ascendente).reset_index(drop=True)

    total = r["VALOR FALTANTES"].sum()
    r.insert(0, "PUESTO", range(1, len(r) + 1))
    r["% DEL COSTO TOTAL"] = (100 * r["VALOR FALTANTES"] / total).round(1) if total else 0.0
    r["% ACUMULADO"] = r["% DEL COSTO TOTAL"].cumsum().round(1)

    return r[[
        "PUESTO", "PICKER", "LINEAS", "TIENDAS CON DIFERENCIA",
        "ENTREGAS CON DIFERENCIA", "ENTREGAS DESPACHADAS", "% ENTREGAS CON DIFERENCIA",
        "CAJAS DESPACHADAS", "MATERIALES DISTINTOS",
        "UNIDADES FALTANTES", "UNIDADES SOBRANTES",
        "VALOR FALTANTES", "VALOR SOBRANTES", "IMPACTO NETO", "IMPACTO ABSOLUTO",
        "COSTO POR CAJA", "COSTO POR ENTREGA DESPACHADA",
        "% DEL COSTO TOTAL", "% ACUMULADO",
    ]]


def picker_por_tienda(datos: Datos) -> pd.DataFrame:
    df = datos.detalle
    r = agrega(df, ["PICKER", "TIENDA"])
    if r.empty:
        return r
    ent = df.groupby(["PICKER", "TIENDA"])["N ENTREGA"].nunique().reset_index(
        name="ENTREGAS CON DIFERENCIA")
    r = r.merge(ent, on=["PICKER", "TIENDA"], how="left")
    return r.sort_values(["VALOR FALTANTES", "PICKER"], ascending=[False, True]).reset_index(drop=True)


def ranking_tiendas(datos: Datos, todas: pd.DataFrame) -> pd.DataFrame:
    r = agrega(todas, ["TIENDA"])
    if r.empty:
        return r
    responsables = (
        datos.detalle.dropna(subset=["PICKER"])
        .groupby(["TIENDA", "PICKER"])["VALOR"]
        .apply(lambda s: s[s < 0].abs().sum())
        .reset_index()
        .sort_values("VALOR", ascending=False)
        .groupby("TIENDA")
        .first()
        .reset_index()
        .rename(columns={"PICKER": "PICKER CON MAYOR IMPACTO",
                         "VALOR": "VALOR DEL PICKER"})
    )
    r = r.merge(responsables, on="TIENDA", how="left")
    r = r.sort_values("VALOR FALTANTES", ascending=False).reset_index(drop=True)
    r.insert(0, "PUESTO", range(1, len(r) + 1))
    return r


def top_materiales(todas: pd.DataFrame, top: int = 100) -> pd.DataFrame:
    r = agrega(todas, ["CODIGO MATERIAL", "DESCRIPCION MATERIAL"])
    if r.empty:
        return r
    return r.sort_values("VALOR FALTANTES", ascending=False).head(top).reset_index(drop=True)


def construye_resumen(datos: Datos, todas: pd.DataFrame, ranking: pd.DataFrame,
                      estado: str, metrica: str) -> pd.DataFrame:
    falt = todas["VALOR"].where(todas["VALOR"] < 0, 0).abs().sum()
    sobr = todas["VALOR"].where(todas["VALOR"] > 0, 0).sum()
    con_p = datos.detalle
    falt_p = con_p["VALOR"].where(con_p["VALOR"] < 0, 0).abs().sum()

    fechas = todas["FECHA"].dropna()
    periodo = (f"{fechas.min():%Y-%m-%d} a {fechas.max():%Y-%m-%d}" if not fechas.empty else "n/d")
    peor = ranking.iloc[0]["PICKER"] if not ranking.empty else "n/d"
    peor_valor = ranking.iloc[0]["VALOR FALTANTES"] if not ranking.empty else 0

    f_des = datos.despachos["FECHA DESPACHO"].dropna()
    periodo_des = (f"{f_des.min():%Y-%m-%d} a {f_des.max():%Y-%m-%d}" if not f_des.empty else "n/d")
    cuadrantes_des = datos.despachos["CUADRANTE DESPACHO"].dropna().unique()
    cuadrantes_des = ", ".join(sorted(cuadrantes_des)) if len(cuadrantes_des) else "n/d"

    filas = [
        ("Periodo analizado", periodo),
        ("Filtro de estado aplicado", estado),
        ("Metrica de ranking", metrica),
        ("Lineas de diferencia analizadas", len(todas)),
        ("Tiendas con diferencias", int(todas["TIENDA"].nunique())),
        ("Materiales con diferencias", int(todas["CODIGO MATERIAL"].nunique())),
        ("Valor faltantes (costo)", falt),
        ("Valor sobrantes", sobr),
        ("Impacto neto", sobr - falt),
        ("Impacto absoluto", falt + sobr),
        ("", ""),
        ("Lineas con picker identificado", datos.cobertura["Lineas con picker identificado"]),
        ("Lineas sin picker identificado", datos.cobertura["Lineas sin picker identificado"]),
        ("% de lineas atribuidas a un picker", datos.cobertura["% lineas atribuidas"]),
        ("Costo atribuido a pickers", falt_p),
        ("Costo sin responsable identificado", falt - falt_p),
        ("Pickers con diferencias", int(con_p["PICKER"].nunique())),
        ("Picker con mayor costo", peor),
        ("Costo del picker con mayor impacto", peor_valor),
        ("", ""),
        ("Periodo cubierto por la hoja de despachos", periodo_des),
        ("Cuadrantes cubiertos por la hoja de despachos", cuadrantes_des),
        ("Entregas en la hoja de despachos", datos.cobertura["Entregas en hoja de despacho"]),
        ("Nota", "Solo se puede asignar picker a las entregas presentes en la hoja "
                 "de despachos; el resto queda en la hoja 'Sin Picker'."),
    ]
    return pd.DataFrame(filas, columns=["INDICADOR", "VALOR"])


# --------------------------------------------------------------------------- #
# Escritura del Excel
# --------------------------------------------------------------------------- #

def formatea_libro(ruta: Path, hojas_dinero: dict[str, list[str]]) -> None:
    from openpyxl import load_workbook
    from openpyxl.chart import BarChart, Reference
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = load_workbook(ruta)
    encabezado = PatternFill("solid", fgColor="1F3864")
    fuente_enc = Font(color="FFFFFF", bold=True)

    for hoja in wb.worksheets:
        if hoja.max_row < 1:
            continue
        titulos = [c.value for c in hoja[1]]
        for celda in hoja[1]:
            celda.fill = encabezado
            celda.font = fuente_enc
            celda.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        hoja.freeze_panes = "A2"
        if hoja.max_row > 1:
            hoja.auto_filter.ref = hoja.dimensions

        columnas_dinero = hojas_dinero.get(hoja.title, [])
        for idx, titulo in enumerate(titulos, start=1):
            letra = get_column_letter(idx)
            largo = max([len(str(titulo))] + [
                len(str(hoja.cell(row=r, column=idx).value or ""))
                for r in range(2, min(hoja.max_row, 300) + 1)
            ])
            hoja.column_dimensions[letra].width = min(max(largo + 2, 11), 46)
            if titulo in columnas_dinero:
                for r in range(2, hoja.max_row + 1):
                    hoja.cell(row=r, column=idx).number_format = FORMATO_PESOS

    # Grafico de los 10 pickers con mayor costo.
    if "Ranking Pickers" in wb.sheetnames:
        hoja = wb["Ranking Pickers"]
        titulos = [c.value for c in hoja[1]]
        if "VALOR FALTANTES" in titulos and hoja.max_row > 1:
            col_valor = titulos.index("VALOR FALTANTES") + 1
            col_picker = titulos.index("PICKER") + 1
            fin = min(hoja.max_row, 11)
            grafico = BarChart()
            grafico.title = "Top 10 pickers por costo de faltantes"
            grafico.y_axis.title = "Pesos"
            grafico.add_data(Reference(hoja, min_col=col_valor, min_row=1, max_row=fin), titles_from_data=True)
            grafico.set_categories(Reference(hoja, min_col=col_picker, min_row=2, max_row=fin))
            grafico.height, grafico.width = 9, 20
            hoja.add_chart(grafico, f"A{hoja.max_row + 3}")

    wb.save(ruta)


def exporta(ruta_salida: Path, tablas: dict[str, pd.DataFrame]) -> None:
    with pd.ExcelWriter(ruta_salida, engine="openpyxl") as writer:
        for nombre, tabla in tablas.items():
            tabla.to_excel(writer, sheet_name=nombre[:31], index=False)

    dinero = ["VALOR", "PRECIO UNITARIO", "VALOR FALTANTES", "VALOR SOBRANTES",
              "IMPACTO NETO", "IMPACTO ABSOLUTO", "COSTO POR CAJA",
              "COSTO POR ENTREGA DESPACHADA", "VALOR DEL PICKER"]
    formatea_libro(ruta_salida, {nombre[:31]: dinero for nombre in tablas})


# --------------------------------------------------------------------------- #
# Salida por consola
# --------------------------------------------------------------------------- #

def pesos(valor) -> str:
    try:
        return f"${valor:,.0f}"
    except (TypeError, ValueError):
        return str(valor)


INDICADORES_EN_PESOS = {
    "Valor faltantes (costo)", "Valor sobrantes", "Impacto neto", "Impacto absoluto",
    "Costo atribuido a pickers", "Costo sin responsable identificado",
    "Costo del picker con mayor impacto",
}


def imprime_consola(resumen: pd.DataFrame, ranking: pd.DataFrame, top: int) -> None:
    print("\n" + "=" * 74)
    print("INFORME DE DIFERENCIAS DE DESPACHO".center(74))
    print("=" * 74)
    for _, fila in resumen.iterrows():
        if fila["INDICADOR"] == "":
            print("-" * 74)
            continue
        valor = fila["VALOR"]
        if fila["INDICADOR"] in INDICADORES_EN_PESOS:
            texto = pesos(valor)
        elif isinstance(valor, (int, float)):
            texto = f"{valor:,}"
        else:
            texto = valor
        print(f"  {fila['INDICADOR']:.<52} {texto}")

    print("\n" + "=" * 74)
    print(f"RANKING DE PICKERS (top {top})".center(74))
    print("=" * 74)
    print(f"  {'#':>2}  {'PICKER':<14} {'LINEAS':>7} {'TIENDAS':>8} "
          f"{'COSTO FALTANTES':>17} {'% TOTAL':>8}")
    print("-" * 74)
    for _, f in ranking.head(top).iterrows():
        print(f"  {f['PUESTO']:>2}  {str(f['PICKER'])[:14]:<14} {f['LINEAS']:>7} "
              f"{f['TIENDAS CON DIFERENCIA']:>8} {pesos(f['VALOR FALTANTES']):>17} "
              f"{f['% DEL COSTO TOTAL']:>7}%")
    print("=" * 74 + "\n")


# --------------------------------------------------------------------------- #
# Programa principal
# --------------------------------------------------------------------------- #

def main(argv=None) -> int:
    p = argparse.ArgumentParser(
        description="Genera un informe detallado de diferencias de despacho por picker.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    p.add_argument("entrada", type=Path, help="Excel con las diferencias (Hoja1) y los despachos (Hoja2).")
    p.add_argument("-o", "--salida", type=Path, default=None,
                   help="Excel de salida (por defecto INFORME_PICKERS_<entrada>.xlsx).")
    p.add_argument("--hoja-diferencias", default="Hoja1", help="Nombre o indice de la hoja de diferencias.")
    p.add_argument("--hoja-despachos", default="Hoja2", help="Nombre o indice de la hoja de despachos.")
    p.add_argument("--estado", default="aplica", choices=["aplica", "todos", "sin-pendientes"],
                   help="'aplica' solo diferencias confirmadas; 'sin-pendientes' excluye las pendientes; "
                        "'todos' no filtra.")
    p.add_argument("--metrica", default="faltante", choices=["faltante", "neto", "absoluto"],
                   help="Criterio para ordenar el ranking de pickers.")
    p.add_argument("--top", type=int, default=15, help="Cuantos pickers mostrar en consola.")
    p.add_argument("--solo-cobertura", action="store_true",
                   help="Analiza unicamente el periodo y los cuadrantes que cubre la hoja de "
                        "despachos, para que los porcentajes por picker sean comparables.")
    p.add_argument("--desde", default=None, help="Fecha inicial AAAA-MM-DD.")
    p.add_argument("--hasta", default=None, help="Fecha final AAAA-MM-DD.")
    args = p.parse_args(argv)

    if not args.entrada.exists():
        print(f"ERROR: no existe el archivo {args.entrada}", file=sys.stderr)
        return 1

    hoja_dif = int(args.hoja_diferencias) if str(args.hoja_diferencias).isdigit() else args.hoja_diferencias
    hoja_des = int(args.hoja_despachos) if str(args.hoja_despachos).isdigit() else args.hoja_despachos

    print(f"Leyendo {args.entrada} ...")
    diferencias = carga_diferencias(args.entrada, hoja_dif)
    despachos = carga_despachos(args.entrada, hoja_des)

    if args.estado == "aplica":
        diferencias = diferencias[diferencias["ESTADO"] == "APLICA"]
    elif args.estado == "sin-pendientes":
        diferencias = diferencias[diferencias["ESTADO"] != "PENDIENTE"]
    if args.solo_cobertura:
        f_des = despachos["FECHA DESPACHO"].dropna()
        if not f_des.empty:
            diferencias = diferencias[
                (diferencias["FECHA"] >= f_des.min())
                & (diferencias["FECHA"] < f_des.max() + pd.Timedelta(days=1))
            ]
        cuadrantes = set(despachos["CUADRANTE DESPACHO"].dropna())
        if cuadrantes:
            diferencias = diferencias[
                diferencias["CUADRANTE"].fillna("").apply(
                    lambda c: any(q in c for q in cuadrantes))
            ]
    if args.desde:
        diferencias = diferencias[diferencias["FECHA"] >= pd.Timestamp(args.desde)]
    if args.hasta:
        diferencias = diferencias[diferencias["FECHA"] <= pd.Timestamp(args.hasta) + pd.Timedelta(days=1)]

    if diferencias.empty:
        print("No quedaron diferencias despues de aplicar los filtros.", file=sys.stderr)
        return 1

    datos = cruza(diferencias, despachos)
    todas = pd.concat([datos.detalle, datos.sin_picker], ignore_index=True)

    ranking = ranking_pickers(datos, args.metrica)
    resumen = construye_resumen(datos, todas, ranking, args.estado, args.metrica)

    columnas_detalle = [
        "FECHA", "PICKER", "TIENDA", "CECO", "CUADRANTE", "N ENTREGA", "RUTA",
        "CODIGO MATERIAL", "DESCRIPCION MATERIAL", "UNIDADES", "PRECIO UNITARIO",
        "VALOR", "TIPO DIFERENCIA", "ESTADO", "CAUSAL", "AREA", "GV", "JDZ", "ATIENDE",
    ]
    detalle = datos.detalle.reindex(columns=columnas_detalle).sort_values(
        ["PICKER", "TIENDA", "VALOR"]).reset_index(drop=True)
    sin_picker = datos.sin_picker.reindex(
        columns=[c for c in columnas_detalle if c not in ("PICKER", "RUTA")]
    ).sort_values("VALOR").reset_index(drop=True)

    salida = args.salida or args.entrada.with_name(f"INFORME_PICKERS_{args.entrada.stem}.xlsx")
    exporta(salida, {
        "Resumen": resumen,
        "Ranking Pickers": ranking,
        "Picker x Tienda": picker_por_tienda(datos),
        "Detalle": detalle,
        "Ranking Tiendas": ranking_tiendas(datos, todas),
        "Top Materiales": top_materiales(todas),
        "Sin Picker": sin_picker,
    })

    imprime_consola(resumen, ranking, args.top)
    print(f"Informe generado: {salida.resolve()}\n")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
